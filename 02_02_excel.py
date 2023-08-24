import openpyxl

path = r"D:\project\This_Is_Crawling\test.xlsx"
wb = openpyxl.Workbook()        # 엑셀 생성
ws = wb.create_sheet('sheet1')  # 시트 생성
ws['A1'] = 'test1'              # 데이터 입력
ws['B1'] = 'abcd'
ws['A2'] = 'test2'
ws['B2'] = 'qwer'
wb.save(path)    # r은 raw data 입력 -> 경로 입력 시 \ 그대로 반영

wb = openpyxl.load_workbook(path)   # 파일 오픈
ws = wb['sheet1']
ws['A3'] = 'test3'
ws['B3'] = 'zxcv'
wb.save(path)
