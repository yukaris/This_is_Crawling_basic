import requests
from bs4 import BeautifulSoup
import openpyxl

codes = ['005930', '000660', '035720']
path = r"D:\project\This_Is_Crawling\stock.xlsx"


def create_excel():
    wb = openpyxl.Workbook()  # 엑셀 생성
    ws = wb.active            # 현재 활성화 시트
    row = 1
    for code in codes:
        response = requests.get(f"https://finance.naver.com/item/sise.naver?code={code}")
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        name = soup.select_one(".wrap_company > h2").text
        ws[f'A{row}'] = name
        row = row + 1
    wb.save(path)


def update_excel():
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    row = 1
    for code in codes:
        response = requests.get(f"https://finance.naver.com/item/sise.naver?code={code}")
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        val = soup.select_one("#_nowVal").text.replace(',', '')
        ws[f'B{row}'] = val
        row = row + 1
    wb.save(path)


create_excel()
update_excel()